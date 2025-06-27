import openpyxl
from collections import defaultdict

# Load workbook and first sheet
file_path = "C:/Users/alexi/OneDrive/Comp sci/ML/Climate Research/SIC_codes.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.worksheets[0]  # First sheet

# Define number ranges and group labels
ranges = {
    "Division B: Mining": (1011, 1499),
    "Division C: Construction": (1521, 1799),
    "Division D: Manufacturing": (2000, 3999),
    "Division E: Transportation, Communications, Electric, Gas, And Sanitary Services": (
        4011,
        4971,
    ),
    "Division F: Wholesale Trade": (5012, 5199),
    "Division G: Retail Trade": (5211, 5999),
    "Division H: Finance, Insurance, And Real Estate": (6011, 6799),
    "Division I: Services": (7011, 8999),
    "Division J: Public Administration": (9111, 9999),
}

# Collect grouped names
grouped_names = defaultdict(list)

for row in range(2, sheet.max_row + 1):
    name = sheet[f"A{row}"].value
    num = sheet[f"B{row}"].value

    if name is None or num is None:
        continue

    try:
        num_value = float(num)
        for label, (low, high) in ranges.items():
            if low <= num_value <= high:
                grouped_names[label].append(str(name))
                break
    except ValueError:
        continue

# Create new sheet for results
if "Grouped Results" in workbook.sheetnames:
    del workbook["Grouped Results"]
output_sheet = workbook.create_sheet("Grouped Results")

# Write results to new sheet
for i, (label, names) in enumerate(grouped_names.items(), start=1):
    output_sheet.cell(row=i, column=1, value=label)
    output_sheet.cell(row=i, column=2, value=len(names))
    output_sheet.cell(row=i, column=3, value=", ".join(names))

# Save the workbook
workbook.save(file_path)
print(
    f"Grouped results written to 'Grouped Results' with counts and names in {file_path}"
)
